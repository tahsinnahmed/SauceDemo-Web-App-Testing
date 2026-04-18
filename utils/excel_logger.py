from openpyxl import Workbook, load_workbook
import os

FILE = "defects.xlsx"

def log_bug(test_name, error):
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Bug ID", "Test Name", "Error", "Status"])
        wb.save(FILE)

    wb = load_workbook(FILE)
    ws = wb.active

    bug_id = f"BUG-{ws.max_row}"

    ws.append([bug_id, test_name, error[:120], "Open"])

    wb.save(FILE)

